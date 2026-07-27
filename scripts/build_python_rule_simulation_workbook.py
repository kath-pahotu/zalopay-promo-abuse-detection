from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Sheet name -> (source CSV written by 01_rule_simulation_storytelling.ipynb,
#                short description of what the sheet is for in Power BI)
CSV_SHEETS = [
    (
        "python_business_rule_simulation_summary.csv",
        "py_rule_summary",
        "Compare rule scenarios",
        "Main chart/table source for Python rule simulation",
    ),
    (
        "python_business_rule_recommendations.csv",
        "py_recommendations",
        "Show recommended action by rule",
        "Use for rule decision table",
    ),
    (
        "python_rule_sensitivity_grid.csv",
        "py_sensitivity_grid",
        "Threshold sensitivity page",
        "Use slicers for rule_family and thresholds",
    ),
    (
        "python_selected_thresholds.csv",
        "py_selected_thresholds",
        "Chosen threshold support",
        "Small summary table",
    ),
    (
        "python_sql_reconciliation_summary.csv",
        "py_sql_reconciliation",
        "QA/reconciliation page",
        "Optional; use to prove Python matches SQL exports",
    ),
]


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing required CSV: {path}")
    return pd.read_csv(path, low_memory=False)


def format_sheet(worksheet, row_count: int, column_count: int) -> None:
    if column_count == 0:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[1].height = 24

    sample_limit = min(row_count + 1, 300)
    for col_idx in range(1, column_count + 1):
        column_letter = get_column_letter(col_idx)
        values = [
            worksheet.cell(row=row_idx, column=col_idx).value
            for row_idx in range(1, sample_limit + 1)
        ]
        max_length = max((len(str(value)) for value in values if value is not None), default=8)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def build_readme_frame() -> pd.DataFrame:
    """Recreates the 'read_me' sheet describing each tab's source CSV and purpose."""
    rows = [
        {
            "Sheet": sheet_name,
            "Source CSV": csv_name,
            "Use in Power BI": use_in_powerbi,
            "Notes": notes,
        }
        for csv_name, sheet_name, use_in_powerbi, notes in CSV_SHEETS
    ]
    return pd.DataFrame(rows)


def build_workbook() -> None:
    # Python notebook writes its CSV outputs to data/output/python_rule_simulation/;
    # the finished workbook goes to data/output/ alongside the SQL exports.
    output_dir = repo_root() / "data" / "output"
    input_dir = output_dir / "python_rule_simulation"
    workbook_path = output_dir / "python_rule_simulation_powerbi.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    summaries = []
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for csv_name, sheet_name, _use_in_powerbi, _notes in CSV_SHEETS:
            csv_path = input_dir / csv_name
            frame = read_csv(csv_path)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.book[sheet_name]
            format_sheet(worksheet, len(frame), len(frame.columns))

            summaries.append((sheet_name, csv_name, len(frame), len(frame.columns)))

        # read_me sheet last, built in-memory (not from a CSV) so it always
        # reflects exactly what's in CSV_SHEETS above.
        readme = build_readme_frame()
        readme.to_excel(writer, sheet_name="read_me", index=False)
        readme_ws = writer.book["read_me"]
        format_sheet(readme_ws, len(readme), len(readme.columns))

    for sheet_name, csv_name, row_count, column_count in summaries:
        print(
            f"sheet name: {sheet_name} | "
            f"source CSV: {csv_name} | "
            f"row count: {row_count} | "
            f"column count: {column_count}"
        )
    print("sheet name: read_me | generated in-memory (not from a CSV)")
    print(f"workbook: {workbook_path}")


if __name__ == "__main__":
    build_workbook()
