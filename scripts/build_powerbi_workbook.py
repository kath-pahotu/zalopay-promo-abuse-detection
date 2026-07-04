from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CSV_SHEETS = [
    ("campaign_overview.csv", "campaign_overview"),
    ("campaign_daily_summary.csv", "campaign_daily_summary"),
    ("campaign_promotion_breakdown.csv", "promotion_breakdown"),
    ("selected_campaign_merchant_distribution.csv", "merchant_distribution"),
    ("selected_campaign_user_scored_features.csv", "user_scored_features"),
    ("suspicious_users_full.csv", "suspicious_users"),
    ("abuse_impact_summary.csv", "abuse_impact"),
    ("rule_simulation_summary.csv", "rule_simulation"),
    ("threshold_percentile_summary.csv", "threshold_percentiles"),
    ("retention_weekly_summary.csv", "weekly_retention"),
]


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing required CSV: {path}")
    return pd.read_csv(path, low_memory=False)


def format_sheet(worksheet, row_count: int, column_count: int) -> None:
    if column_count == 0:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[1].height = 24

    sample_limit = min(row_count + 1, 300)
    for col_idx in range(1, column_count + 1):
        column_letter = get_column_letter(col_idx)
        values = [
            worksheet.cell(row=row_idx, column=col_idx).value
            for row_idx in range(1, sample_limit + 1)
        ]
        max_length = max((len(str(value)) for value in values if value is not None), default=8)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def build_workbook() -> None:
    output_dir = repo_root() / "data" / "output"
    workbook_path = output_dir / "powerbi_outputs.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    summaries = []
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for csv_name, sheet_name in CSV_SHEETS:
            csv_path = output_dir / csv_name
            frame = read_csv(csv_path)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.book[sheet_name]
            format_sheet(worksheet, len(frame), len(frame.columns))

            summaries.append((sheet_name, csv_name, len(frame), len(frame.columns)))

    for sheet_name, csv_name, row_count, column_count in summaries:
        print(
            f"sheet name: {sheet_name} | "
            f"source CSV: {csv_name} | "
            f"row count: {row_count} | "
            f"column count: {column_count}"
        )

    print(f"workbook: {workbook_path}")


if __name__ == "__main__":
    build_workbook()
